from openpyxl import load_workbook
import csv
import openpyxl

wb = load_workbook('memberships.xlsx')
sheet = wb['group membership']
with open('out.csv', 'w', newline='') as out_file:
    writer = csv.writer(out_file, delimiter=',')
    i = 1
    while i <= sheet.max_row:
        test = sheet.cell(row=i, column=2).value
        if (sheet.cell(row=i, column=2).value.find('ITS-G') != -1):
            strToWrite = sheet.cell(row=i, column=1).value + ',' + sheet.cell(row=i, column=2).value
            writer.writerow([strToWrite])
            i += 1
        else:
            i += 1
